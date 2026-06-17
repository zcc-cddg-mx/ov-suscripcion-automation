"""
Generator for migration type 2: rating rules (VHPlanRules, VHPlanSetup, etc.) in ams-rule.

Reads the current version from the last matching migration in the target repo,
increments it, and produces a Flyway-ready Excel with sheets:
  RuleKit | RatingList | <EntityName> | LOV
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_FIXTURES = Path(__file__).parent.parent / "fixtures"
_LOV_FILE = _FIXTURES / "lov_ams_rule.json"

_RULEKIT_HEADERS = ["ID", "Code", "Name", "Description", "Version", "State"]

_RATINGLIST_HEADERS = [
    "ID", "Rule kit", "Type", "Code", "Name", "Description",
    "Selector", "Selector type", "Version", "Valid from", "Valid to", "State",
]


def _entity_to_code(entity_name: str) -> str:
    """VHPlanRules → VH_PLAN_RULES.

    Inserts _ between:
    - a lowercase letter and an uppercase letter  (Plan → _Plan)
    - a run of uppercase letters and the start of the next word (VHP → VH_P)
    """
    # Between a run of uppercase and an upcoming uppercase+lowercase  (e.g. VHPlan → VH_Plan)
    result = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", "_", entity_name)
    # Between a lowercase letter and an uppercase letter (e.g. PlanRules → Plan_Rules)
    result = re.sub(r"(?<=[a-z])(?=[A-Z])", "_", result)
    return result.upper()


def _find_last_migration(repo_resources_path: Path, entity_name: str) -> Path | None:
    """Return the most recent xlsx for *entity_name* in the migration resources directory."""
    pattern = f"*_{entity_name}.xlsx"
    candidates = sorted(repo_resources_path.glob(pattern))
    return candidates[-1] if candidates else None


def _read_current_version(xlsx_path: Path) -> tuple[int, datetime | None]:
    """
    From a previous migration xlsx, read the NEW row in RatingList.
    Returns (version, valid_from_date).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RatingList"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    id_idx = list(headers).index("ID")
    ver_idx = list(headers).index("Version")
    from_idx = list(headers).index("Valid from")

    for row in rows[1:]:
        if row[id_idx] == "NEW":
            version = row[ver_idx]
            valid_from = row[from_idx]
            if isinstance(valid_from, str):
                # formula string — use today as fallback
                valid_from = None
            return int(version), valid_from

    raise ValueError(f"No NEW row found in RatingList of {xlsx_path}")


def _load_raw_rules(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Return (headers, data_rows) from the raw input Excel."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty workbook: {path}")
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data = [list(row) for row in rows[1:] if any(c is not None for c in row)]
    return headers, data


_COTIZADOR_FACTOR_COLS = [
    "DAPA_FREC", "DAPA_CM", "DATO_FREC", "DATO_CM",
    "RC_FREC", "RC_CM", "ROPA_FREC", "ROPA_CM", "ROTO_FREC", "ROTO_CM",
]

_COTIZADOR_DRIVERS_AGE_HEADERS = [
    "Drivers age from", "Drivers age to",
] + _COTIZADOR_FACTOR_COLS


def _open_workbook(path: Path, password: str | None = None) -> openpyxl.Workbook:
    """Open an xlsx, decrypting with *password* if provided."""
    if password:
        import msoffcrypto
        with open(path, "rb") as fh:
            office = msoffcrypto.OfficeFile(fh)
            office.load_key(password=password)
            buf = io.BytesIO()
            office.decrypt(buf)
        return openpyxl.load_workbook(buf, data_only=True)
    return openpyxl.load_workbook(path, data_only=True)


def _load_cotizador_drivers_age(
    path: Path, password: str | None = None
) -> tuple[list[str], list[list[Any]]]:
    """
    Extract VHDriversAge rows from the cotizador Relatividades sheet.

    Locates the EDAD_INPUT table (header row where col B = 'EDAD_INPUT'),
    reads factor columns C–L (10 columns), and applies:
      - ages '17'–'79'  → from=int(age), to=int(age)
      - age  '80+'      → from=80, to=998
      - 'P.JURIDICA'    → skipped (legal entity, not a person age)
      - sentinel row    → from=999, to=999, all factors=1  (always appended)
    """
    wb = _open_workbook(path, password)
    ws = wb["Relatividades"]
    rows = list(ws.iter_rows(values_only=True))

    # Find EDAD_INPUT header row (col index 1 = B)
    start = None
    for i, row in enumerate(rows):
        if row[1] == "EDAD_INPUT":
            start = i + 1  # data starts on next row
            break
    if start is None:
        raise ValueError(f"EDAD_INPUT table not found in 'Relatividades' sheet of {path}")

    data: list[list[Any]] = []
    for row in rows[start:]:
        age_val = row[1]
        if age_val is None:
            break  # end of table (blank row)
        age_str = str(age_val).strip()
        if age_str == "P.JURIDICA":
            continue  # skip — legal entity row, not an age

        factors = list(row[2:12])  # columns C–L (10 factor columns)

        if age_str == "80+":
            age_from, age_to = 80, 998
        else:
            try:
                age_int = int(age_str)
            except ValueError:
                raise ValueError(
                    f"Unexpected age value '{age_str}' in EDAD_INPUT table of {path}"
                )
            age_from = age_int
            age_to = age_int

        data.append([age_from, age_to] + factors)

    if not data:
        raise ValueError(f"No data rows found in EDAD_INPUT table of {path}")

    # Append sentinel row (999–999, all factors = 1)
    data.append([999, 999] + [1] * 10)

    return _COTIZADOR_DRIVERS_AGE_HEADERS, data


def _write_rulekit(ws: Worksheet) -> None:
    ws.append(_RULEKIT_HEADERS)


def _write_ratinglist(
    ws: Worksheet,
    entity_code: str,
    entity_name: str,
    old_version: int,
    old_valid_from: datetime | None,
) -> None:
    ws.append(_RATINGLIST_HEADERS)
    new_version = old_version + 1

    old_from_val: Any = old_valid_from if old_valid_from else datetime(2024, 1, 1)

    ws.append([
        "OLD", None, "Global", entity_code, entity_name, entity_name,
        None, "None", old_version, old_from_val, "=TODAY()-(0.5/24)", "Draft",
    ])
    ws.append([
        "NEW", None, "Global", entity_code, entity_name, entity_name,
        None, "None", new_version, "=TODAY()", None, "Draft",
    ])


def _write_entity_sheet(
    ws: Worksheet,
    raw_headers: list[str],
    raw_data: list[list[Any]],
) -> None:
    """Write headers then data rows, prepending None (ID) and 'NEW' (Rating list)."""
    full_headers = ["ID", "Rating list"] + raw_headers
    ws.append(full_headers)
    for row in raw_data:
        ws.append([None, "NEW"] + row)


def _write_lov(ws: Worksheet) -> None:
    lov_data: list[list[Any]] = json.loads(_LOV_FILE.read_text(encoding="utf-8"))
    for row in lov_data:
        ws.append(row)


_COTIZADOR_LOADERS: dict[str, Any] = {
    "VHDriversAge": _load_cotizador_drivers_age,
}


def generate(
    raw_input: Path,
    output: Path,
    entity_name: str,
    repo_resources_path: Path,
    password: str | None = None,
) -> None:
    """
    Transform *raw_input* into a Flyway-ready rules Excel at *output*.

    *entity_name*: e.g. "VHPlanRules", "VHDriversAge"
    *repo_resources_path*: path to ams-rule/flyway/src/main/resources/db/migration/
    *password*: Excel password if the input file is encrypted (e.g. cotizador files)
    """
    entity_code = _entity_to_code(entity_name)

    loader = _COTIZADOR_LOADERS.get(entity_name)
    if loader is not None:
        raw_headers, raw_data = loader(raw_input, password)
    else:
        raw_headers, raw_data = _load_raw_rules(raw_input)

    last_migration = _find_last_migration(repo_resources_path, entity_name)
    if last_migration is None:
        raise FileNotFoundError(
            f"No previous migration found for {entity_name} in {repo_resources_path}"
        )
    current_version, valid_from = _read_current_version(last_migration)

    wb = Workbook()

    ws_rulekit = wb.active
    ws_rulekit.title = "RuleKit"
    _write_rulekit(ws_rulekit)

    ws_rating = wb.create_sheet("RatingList")
    _write_ratinglist(ws_rating, entity_code, entity_name, current_version, valid_from)

    ws_entity = wb.create_sheet(entity_name)
    _write_entity_sheet(ws_entity, raw_headers, raw_data)

    ws_lov = wb.create_sheet("LOV")
    _write_lov(ws_lov)

    wb.save(output)
