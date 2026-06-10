"""
Generator for migration type 1: VH_ren_data (ams-policy).

Transforms a raw business Excel into the Flyway-ready Excel with sheets: LOV, FixedRenewalData.

Expected input (real files from negocio):
  - Columns: CHASIS, TASA FINAL, PLACAS  (+ possible empty trailing columns)
  - Year and Month are NOT in the file — must be passed as arguments.
  - Factor rules:
      · Numeric: must have at most 8 decimal places. Values with more are
        truncated to 8 (this is normal — negocio files often carry 10–18 decimals).
      · 'No Renovar' (string): business rule — row is included with
        Factor='No Renovar' and Renewal blocked='Yes'.
      · Any other non-numeric, non-None value raises ValueError.
  - The file contains empty trailing rows that are filtered automatically.
"""

from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_FIXTURES = Path(__file__).parent.parent / "fixtures"
_LOV_FILE = _FIXTURES / "lov_ams_policy.json"

# Accepted column name aliases → canonical name
_COL_ALIASES: dict[str, str] = {
    "chassis": "Chassis number",
    "chassis number": "Chassis number",
    "chasis": "Chassis number",
    "year": "Year",
    "año": "Year",
    "anio": "Year",
    "month": "Month",
    "mes": "Month",
    "factor": "Factor",
    "tasa final": "Factor",
    "tasa": "Factor",
    "sum insured": "Sum insured",
    "suma asegurada": "Sum insured",
    "suma_asegurada": "Sum insured",
    "renewal blocked": "Renewal blocked",
    "bloqueado": "Renewal blocked",
}

_FIXED_HEADERS = [
    "FixedRenewalData",
    "Year",
    "Month",
    "Chassis number",
    "Sum insured",
    "Factor",
    "Renewal blocked",
]

_NO_RENOVAR = "no renovar"
_MAX_FACTOR_DECIMALS = 8


class _RowErrors:
    def __init__(self) -> None:
        self._messages: list[str] = []

    def add(self, msg: str) -> None:
        self._messages.append(msg)

    def has_errors(self) -> bool:
        return bool(self._messages)

    def raise_if_any(self) -> None:
        if self._messages:
            detail = "\n".join(f"  • {m}" for m in self._messages)
            raise ValueError(
                f"Validation failed: {len(self._messages)} error(s) in input file:\n{detail}"
            )


def _normalize_header(name: str) -> str:
    return _COL_ALIASES.get(name.strip().lower(), name.strip())


def _is_no_renovar(value: Any) -> bool:
    return isinstance(value, str) and value.strip().lower() == _NO_RENOVAR


def _validate_chassis(value: Any) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "Empty chassis number"
    return None


def _decimal_places(value: float) -> int:
    """Return the number of significant decimal places of *value*.

    Uses Decimal(str(value)) to avoid float binary-representation noise.
    e.g. 0.01963615 → 8, not 18.
    """
    d = Decimal(str(value)).normalize()
    sign, digits, exponent = d.as_tuple()
    if exponent >= 0:
        return 0
    return -exponent


def _validate_and_normalize_factor(value: Any, chassis: str, row_num: int) -> Any:
    """
    Validate and normalize the Factor value for a single row.

    - 'No Renovar' (case-insensitive): returned as-is (handled by caller).
    - Numeric (int/float): truncated to _MAX_FACTOR_DECIMALS decimals if needed.
    - Anything else: raises ValueError.
    """
    if _is_no_renovar(value):
        return value

    if isinstance(value, int):
        return float(value)

    if isinstance(value, float):
        decimals = _decimal_places(value)
        if decimals > _MAX_FACTOR_DECIMALS:
            return round(value, _MAX_FACTOR_DECIMALS)
        return value

    if value is None:
        msg = "Factor is empty"
    elif isinstance(value, str) and not value.strip():
        msg = "Factor is empty (whitespace)"
    else:
        msg = f"Factor must be numeric or 'No Renovar', got {value!r} ({type(value).__name__})"
    raise ValueError(f"Row {row_num} (chassis {chassis!r}): {msg}")


def _load_raw(path: Path, year: int, month: int) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty workbook: {path}")

    raw_headers = [str(c) if c is not None else "" for c in rows[0]]
    headers = [_normalize_header(h) for h in raw_headers]

    has_year = "Year" in headers
    has_month = "Month" in headers

    required = {"Chassis number", "Factor"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Missing required columns in input: {missing}")

    errors = _RowErrors()
    record_with_rows: list[tuple[dict[str, Any], int]] = []
    count_no_renovar = 0
    count_normalized = 0

    for row_num, raw_row in enumerate(rows[1:], start=2):
        if not any(c is not None for c in raw_row):
            continue  # skip empty trailing rows

        rec = dict(zip(headers, raw_row))
        chassis = rec.get("Chassis number")
        raw_factor = rec.get("Factor")

        chassis_err = _validate_chassis(chassis)
        if chassis_err:
            errors.add(f"Row {row_num}: {chassis_err}")
            continue

        try:
            normalized = _validate_and_normalize_factor(raw_factor, chassis, row_num)
        except ValueError as exc:
            errors.add(str(exc))
            continue

        if _is_no_renovar(normalized):
            rec["Factor"] = normalized
            rec["Renewal blocked"] = "Yes"
            count_no_renovar += 1
        else:
            if normalized != raw_factor:
                count_normalized += 1
            rec["Factor"] = normalized

        if not has_year:
            rec["Year"] = year
        if not has_month:
            rec["Month"] = month

        record_with_rows.append((rec, row_num))

    # Detect duplicate chassis numbers
    seen: dict[str, int] = {}
    for rec, row_num in record_with_rows:
        ch = str(rec.get("Chassis number", ""))
        if ch in seen:
            errors.add(
                f"Row {row_num} (chassis {ch!r}): Duplicate chassis — first seen at row {seen[ch]}"
            )
        else:
            seen[ch] = row_num

    errors.raise_if_any()

    records = [rec for rec, _ in record_with_rows]

    if count_no_renovar:
        print(f"  {count_no_renovar} rows with 'No Renovar' included (Renewal blocked=Yes)")
    if count_normalized:
        print(f"  {count_normalized} Factor values normalized to {_MAX_FACTOR_DECIMALS} decimal places")

    # Sort: numeric factors ascending, 'No Renovar' rows at the end
    records.sort(key=lambda r: (
        1 if _is_no_renovar(r.get("Factor")) else 0,
        r.get("Factor") if not _is_no_renovar(r.get("Factor")) else 0,
    ))

    return records


def _write_lov(ws: Worksheet) -> None:
    lov_data: list[list[Any]] = json.loads(_LOV_FILE.read_text(encoding="utf-8"))
    for row in lov_data:
        ws.append(row)


def _write_fixed_renewal(ws: Worksheet, records: list[dict[str, Any]]) -> None:
    ws.append(_FIXED_HEADERS)
    for rec in records:
        ws.append([
            None,
            rec.get("Year"),
            rec.get("Month"),
            rec.get("Chassis number"),
            rec.get("Sum insured"),
            rec.get("Factor"),
            rec.get("Renewal blocked", "No"),
        ])


def generate(raw_input: Path, output: Path, year: int, month: int) -> None:
    """
    Transform *raw_input* Excel into a Flyway-ready ren_data Excel at *output*.

    *year* and *month* are required because real business files from negocio
    do not include those columns — only CHASIS and TASA FINAL.
    """
    records = _load_raw(raw_input, year, month)
    if not records:
        raise ValueError(
            "No valid records found in input file — all rows were empty or failed validation."
        )

    wb = Workbook()
    ws_lov = wb.active
    ws_lov.title = "LOV"
    _write_lov(ws_lov)

    ws_data = wb.create_sheet("FixedRenewalData")
    _write_fixed_renewal(ws_data, records)

    wb.save(output)
    print(f"  {len(records)} records written to {output.name}")
