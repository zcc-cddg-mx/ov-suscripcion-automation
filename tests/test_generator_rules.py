"""Tests for generator_rules."""

import os
from pathlib import Path

import openpyxl
import pytest

from src.generator_rules import (
    generate, _entity_to_code, _read_current_version,
    _load_cotizador_drivers_age, _COTIZADOR_DRIVERS_AGE_HEADERS,
)

_FIXTURES = Path(__file__).parent.parent / "fixtures"
_RULES_REF = _FIXTURES / "VHPlanRules_reference.xlsx"
_COTIZADOR = _FIXTURES / "rules/drivers-age/Cotizador_MotorIndividualV25- Modelos Enero2026.xlsx"
_COTIZADOR_PASSWORD = os.environ.get("BUSINESS_EXCEL_PASSWORD", "Motor2023*")

# Path to a real ams-rule migration directory (relative to this file's parent parent parent)
_AMS_RULE_RESOURCES = (
    Path(__file__).parent.parent.parent
    / "ov-arizona-backend-ecuador"
    / "ams-rule/flyway/src/main/resources/db/migration"
)


def _make_raw_rules_input(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Transaction type", "Plan type", "Sub product", "Role type",
        "Years from", "Years to", "Factor", "Factor min", "Factor max",
        "Days to renewal", "Ensurance line of business", "Premium min",
        "Owner civil status code", "Owner drivers age from", "Owner drivers age to",
        "Sum insured from", "Sum insured to",
    ]
    ws.append(headers)
    ws.append(["NEWBUSINESS", "Beginner", "MODULAR", "UNW", 0, 10, 0.5, None, None, 45, "11793261264993", None, None, None, None, 0, 999999])
    wb.save(path)


class TestEntityToCode:
    def test_vh_plan_rules(self) -> None:
        assert _entity_to_code("VHPlanRules") == "VH_PLAN_RULES"

    def test_vh_plan_setup(self) -> None:
        assert _entity_to_code("VHPlanSetup") == "VH_PLAN_SETUP"

    def test_single_word(self) -> None:
        assert _entity_to_code("Factor") == "FACTOR"


class TestReadCurrentVersion:
    def test_reads_version_from_reference(self) -> None:
        version, valid_from = _read_current_version(_RULES_REF)
        # The reference file has NEW version = 19
        assert version == 19

    def test_no_new_row_raises(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RatingList"
        ws.append(["ID", "Rule kit", "Type", "Code", "Name", "Description",
                   "Selector", "Selector type", "Version", "Valid from", "Valid to", "State"])
        ws.append(["OLD", None, "Global", "CODE", "Name", "Name", None, "None", 5, None, None, "Draft"])
        out = tmp_path / "no_new.xlsx"
        wb.save(out)
        with pytest.raises(ValueError, match="No NEW row"):
            _read_current_version(out)


class TestGeneratorRules:
    def test_output_sheets(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        generate(raw, out, "VHPlanRules", _AMS_RULE_RESOURCES)

        wb = openpyxl.load_workbook(out, data_only=True)
        assert wb.sheetnames == ["RuleKit", "RatingList", "VHPlanRules", "LOV"]

    def test_ratinglist_version_incremented(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        generate(raw, out, "VHPlanRules", _AMS_RULE_RESOURCES)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["RatingList"]
        rows = list(ws.iter_rows(values_only=True))
        # rows[0] = headers, rows[1] = OLD, rows[2] = NEW
        old_version = rows[1][8]
        new_version = rows[2][8]
        assert new_version == old_version + 1

    def test_ratinglist_has_today_formula(self, tmp_path: Path) -> None:
        """Valid from NEW must be the =TODAY() formula string."""
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        generate(raw, out, "VHPlanRules", _AMS_RULE_RESOURCES)

        # Load without data_only to read formula strings
        wb = openpyxl.load_workbook(out, data_only=False)
        ws = wb["RatingList"]
        # Row 3 (index 2), col J (index 10, 1-based = 10)
        new_valid_from = ws.cell(3, 10).value  # Valid from NEW
        assert new_valid_from == "=TODAY()"

    def test_ratinglist_old_valid_to_formula(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        generate(raw, out, "VHPlanRules", _AMS_RULE_RESOURCES)

        wb = openpyxl.load_workbook(out, data_only=False)
        ws = wb["RatingList"]
        old_valid_to = ws.cell(2, 11).value  # Valid to OLD
        assert old_valid_to == "=TODAY()-(0.5/24)"

    def test_entity_sheet_has_id_and_rating_list_cols(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        generate(raw, out, "VHPlanRules", _AMS_RULE_RESOURCES)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["VHPlanRules"]
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Rating list"
        assert ws.cell(2, 1).value is None      # ID always None
        assert ws.cell(2, 2).value == "NEW"

    def test_lov_is_non_empty(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        generate(raw, out, "VHPlanRules", _AMS_RULE_RESOURCES)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["LOV"]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        assert len(rows) > 0
        # First row must be TrueFalse entry
        assert rows[0][0] == "TrueFalse"

    def test_missing_previous_migration_raises(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_rules_input(raw)
        out = tmp_path / "output.xlsx"
        empty_dir = tmp_path / "empty_migration"
        empty_dir.mkdir()
        with pytest.raises(FileNotFoundError):
            generate(raw, out, "VHPlanRules", empty_dir)


class TestCotizadorDriversAge:
    """Tests for _load_cotizador_drivers_age — extracts EDAD_INPUT from the cotizador."""

    def test_headers(self) -> None:
        headers, _ = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        assert headers == _COTIZADOR_DRIVERS_AGE_HEADERS

    def test_row_count(self) -> None:
        _, data = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        # 63 single-year ages (17–79) + 1 for '80+' + 1 sentinel = 65
        assert len(data) == 65

    def test_first_row_age_17(self) -> None:
        _, data = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        assert data[0][0] == 17  # from
        assert data[0][1] == 17  # to

    def test_80plus_maps_to_80_998(self) -> None:
        _, data = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        # Second-to-last row is 80+ mapping
        assert data[-2][0] == 80
        assert data[-2][1] == 998

    def test_sentinel_row_last(self) -> None:
        _, data = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        assert data[-1][0] == 999
        assert data[-1][1] == 999
        assert all(v == 1 for v in data[-1][2:])

    def test_pjuridica_excluded(self) -> None:
        _, data = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        ages = [r[0] for r in data]
        assert "P.JURIDICA" not in ages
        assert None not in ages

    def test_each_row_has_10_factor_columns(self) -> None:
        _, data = _load_cotizador_drivers_age(_COTIZADOR, _COTIZADOR_PASSWORD)
        for row in data:
            assert len(row) == 12  # 2 age cols + 10 factors

    def test_generate_full_pipeline(self, tmp_path: Path) -> None:
        """End-to-end: cotizador → VHDriversAge migration xlsx with correct structure."""
        out = tmp_path / "VHDriversAge_test.xlsx"
        generate(
            raw_input=_COTIZADOR,
            output=out,
            entity_name="VHDriversAge",
            repo_resources_path=_AMS_RULE_RESOURCES,
            password=_COTIZADOR_PASSWORD,
        )
        wb = openpyxl.load_workbook(out, data_only=True)
        assert wb.sheetnames == ["RuleKit", "RatingList", "VHDriversAge", "LOV"]

        ws = wb["VHDriversAge"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == tuple(["ID", "Rating list"] + _COTIZADOR_DRIVERS_AGE_HEADERS)
        assert len(rows) - 1 == 65  # 65 data rows

        # Version must be auto-incremented
        ws_rl = wb["RatingList"]
        rl = list(ws_rl.iter_rows(values_only=True))
        assert rl[2][8] == rl[1][8] + 1  # NEW version = OLD + 1
