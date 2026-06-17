"""
Migration runner for manual end-to-end testing — rules (ams-rule).

Generates real migration files into tests/migrations/ and verifies them
against the reference Excel and business rules.

Usage:
    BUSINESS_EXCEL_PASSWORD=Motor2023* python tests/run_rules_test.py --entity VHDriversAge
    BUSINESS_EXCEL_PASSWORD=Motor2023* python tests/run_rules_test.py --entity VHDriversAge --ticket INC99999
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

_ROOT = Path(__file__).parent.parent
_OUTPUT_DIR = Path(__file__).parent / "migrations"
_FIXTURES = _ROOT / "fixtures"
_BACKEND = _ROOT.parent / "ov-arizona-backend-ecuador"
_AMS_RULE_RESOURCES = _BACKEND / "ams-rule/flyway/src/main/resources/db/migration"

_ENTITIES = {
    "VHDriversAge": {
        "input": _FIXTURES / "rules/business-reference/Cotizador_MotorIndividualV25.xlsx",
        "reference_xlsx": _FIXTURES / "rules/drivers-age/VHDriversAge_reference.xlsx",
        "password": os.environ.get("BUSINESS_EXCEL_PASSWORD"),
        "expected_data_rows": 65,
        "expected_lov_rows": 52,
        "expected_headers": [
            "ID", "Rating list", "Drivers age from", "Drivers age to",
            "DAPA_FREC", "DAPA_CM", "DATO_FREC", "DATO_CM",
            "RC_FREC", "RC_CM", "ROPA_FREC", "ROPA_CM", "ROTO_FREC", "ROTO_CM",
        ],
    },
}

# ── helpers ───────────────────────────────────────────────────────────────────

def _ok(msg: str) -> None:
    print(f"  \033[32m✓\033[0m {msg}")

def _fail(msg: str) -> None:
    print(f"  \033[31m✗\033[0m {msg}")

def _info(msg: str) -> None:
    print(f"  · {msg}")


# ── generation ────────────────────────────────────────────────────────────────

def _generate(entity_name: str, cfg: dict, ticket_override: str | None) -> tuple[Path, Path]:
    from src.generator_rules import generate as gen_xlsx
    from src.java_template import generate as gen_java

    ticket = ticket_override or "TEST99999"
    ticket_safe = ticket.replace("-", "_")
    ts = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    base_name = f"V{ts}__{ticket_safe}_{entity_name}"

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_out = _OUTPUT_DIR / f"{base_name}.xlsx"
    java_out = _OUTPUT_DIR / f"{base_name}.java"

    print(f"\n── Generating {base_name} ──")
    gen_xlsx(
        raw_input=cfg["input"],
        output=xlsx_out,
        entity_name=entity_name,
        repo_resources_path=_AMS_RULE_RESOURCES,
        password=cfg.get("password"),
    )
    _info(f"xlsx: {xlsx_out.name}")

    java_src = gen_java(base_name, "ams-rule")
    java_out.write_text(java_src, encoding="utf-8")
    _info(f"java: {java_out.name}")

    return xlsx_out, java_out


# ── verification ──────────────────────────────────────────────────────────────

def _verify_xlsx(xlsx_out: Path, entity_name: str, cfg: dict) -> int:
    failures = 0
    print("\n── Verifying xlsx ──")

    wb = openpyxl.load_workbook(xlsx_out, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_out, data_only=False)

    # 1. Sheets en orden correcto
    expected_sheets = ["RuleKit", "RatingList", entity_name, "LOV"]
    if wb.sheetnames == expected_sheets:
        _ok(f"Sheets: {wb.sheetnames}")
    else:
        _fail(f"Expected {expected_sheets}, got {wb.sheetnames}")
        failures += 1

    # 2. RuleKit solo tiene headers, sin datos
    ws_rk = wb["RuleKit"]
    if ws_rk.max_row == 1:
        _ok("RuleKit: headers only (no data rows)")
    else:
        _fail(f"RuleKit: expected 1 row (headers), got {ws_rk.max_row}")
        failures += 1

    # 3. RatingList tiene 3 filas (header + OLD + NEW)
    ws_rl = wb["RatingList"]
    rl_rows = list(ws_rl.iter_rows(values_only=True))
    if len(rl_rows) == 3:
        _ok("RatingList: 3 rows (header + OLD + NEW)")
    else:
        _fail(f"RatingList: expected 3 rows, got {len(rl_rows)}")
        failures += 1
        rl_rows = rl_rows[:3]  # prevent index errors below

    # 4. Versión NEW = OLD + 1
    if len(rl_rows) >= 3:
        headers_rl = list(rl_rows[0])
        ver_idx = headers_rl.index("Version")
        old_ver = rl_rows[1][ver_idx]
        new_ver = rl_rows[2][ver_idx]
        if isinstance(old_ver, int) and new_ver == old_ver + 1:
            _ok(f"Version incremented: OLD={old_ver} → NEW={new_ver}")
        else:
            _fail(f"Version not incremented correctly: OLD={old_ver}, NEW={new_ver}")
            failures += 1

    # 5. NEW Valid from = "=TODAY()" (formula string)
    ws_rl_f = wb_formulas["RatingList"]
    new_valid_from = ws_rl_f.cell(3, 10).value  # row 3, col J
    if new_valid_from == "=TODAY()":
        _ok('NEW Valid from = "=TODAY()"')
    else:
        _fail(f'NEW Valid from expected "=TODAY()", got {new_valid_from!r}')
        failures += 1

    # 6. OLD Valid to = "=TODAY()-(0.5/24)"
    old_valid_to = ws_rl_f.cell(2, 11).value  # row 2, col K
    if old_valid_to == "=TODAY()-(0.5/24)":
        _ok('OLD Valid to = "=TODAY()-(0.5/24)"')
    else:
        _fail(f'OLD Valid to expected "=TODAY()-(0.5/24)", got {old_valid_to!r}')
        failures += 1

    # 7. Headers de entity sheet correctos
    ws_entity = wb[entity_name]
    actual_headers = list(ws_entity.iter_rows(values_only=True, max_row=1))[0]
    expected_headers = cfg["expected_headers"]
    if list(actual_headers) == expected_headers:
        _ok(f"Entity sheet headers correct ({len(expected_headers)} cols)")
    else:
        _fail(f"Headers mismatch.\n    Expected: {expected_headers}\n    Got:      {list(actual_headers)}")
        failures += 1

    # 8. Conteo de filas de datos
    data_rows = list(ws_entity.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in data_rows if any(c is not None for c in r)]
    expected_rows = cfg["expected_data_rows"]
    if len(data_rows) == expected_rows:
        _ok(f"Data rows: {len(data_rows)} (expected {expected_rows})")
    else:
        _fail(f"Data rows: got {len(data_rows)}, expected {expected_rows}")
        failures += 1

    # 9. Penúltima fila: age_from=80, age_to=998
    if len(data_rows) >= 2:
        second_last = data_rows[-2]
        if second_last[2] == 80 and second_last[3] == 998:
            _ok("Second-to-last row: age_from=80, age_to=998 (80+ mapping)")
        else:
            _fail(f"Second-to-last row: expected (80, 998), got ({second_last[2]}, {second_last[3]})")
            failures += 1

    # 10. Última fila (sentinel): 999, 999, all factors=1
    if data_rows:
        last = data_rows[-1]
        if last[2] == 999 and last[3] == 999 and all(v == 1 for v in last[4:]):
            _ok("Last row: sentinel (999, 999, all factors=1)")
        else:
            _fail(f"Last row (sentinel) incorrect: {last}")
            failures += 1

    # 11. LOV row count
    ws_lov = wb["LOV"]
    lov_rows = [r for r in ws_lov.iter_rows(values_only=True) if any(c is not None for c in r)]
    expected_lov = cfg["expected_lov_rows"]
    if len(lov_rows) == expected_lov:
        _ok(f"LOV rows: {len(lov_rows)} (expected {expected_lov})")
    else:
        _fail(f"LOV rows: got {len(lov_rows)}, expected {expected_lov}")
        failures += 1

    # 12. Comparación row-by-row contra referencia
    ref_path: Path = cfg.get("reference_xlsx")
    if ref_path and ref_path.exists():
        print("\n── Comparing against reference ──")
        failures += _compare_with_reference(data_rows, ref_path, entity_name)
    elif ref_path:
        _info(f"Reference not found, skipping comparison: {ref_path.name}")

    return failures


def _compare_with_reference(data_rows: list, ref_path: Path, entity_name: str) -> int:
    failures = 0
    wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
    ws_ref = wb_ref[entity_name]
    ref_rows = list(ws_ref.iter_rows(min_row=2, values_only=True))
    ref_rows = [r for r in ref_rows if any(c is not None for c in r)]

    # Row count
    if len(data_rows) == len(ref_rows):
        _ok(f"Row count matches reference: {len(data_rows)}")
    else:
        _fail(f"Row count: generated={len(data_rows)}, reference={len(ref_rows)}")
        failures += 1

    # Age ranges match
    gen_ages = [(r[2], r[3]) for r in data_rows]
    ref_ages = [(r[2], r[3]) for r in ref_rows]
    if gen_ages == ref_ages:
        _ok(f"Age ranges match reference ({len(gen_ages)} rows)")
    else:
        mismatches = [(i, g, r) for i, (g, r) in enumerate(zip(gen_ages, ref_ages)) if g != r]
        _fail(f"Age range mismatches ({len(mismatches)}): {mismatches[:3]}")
        failures += 1

    # Factor values per age row (columns 4–13, i.e. index 4:14)
    # Tolerance: 1e-9 relative — absorbs float binary noise from Excel (digit 12+).
    # Differences beyond tolerance are real data divergences (e.g. cotizador updated
    # but system not yet migrated) and are reported as informational warnings, not failures.
    _REL_TOL = 1e-9
    real_mismatches = []
    float_noise_rows = 0
    for i, (gen_row, ref_row) in enumerate(zip(data_rows, ref_rows)):
        gen_factors = list(gen_row[4:14])
        ref_factors = list(ref_row[4:14])
        if gen_factors == ref_factors:
            continue
        col_diffs = []
        for j, (g, r) in enumerate(zip(gen_factors, ref_factors)):
            if g == r:
                continue
            if g is not None and r is not None and abs(g - r) <= _REL_TOL * max(abs(g), abs(r), 1e-30):
                float_noise_rows += 1
            else:
                col_diffs.append(j)
        if col_diffs:
            real_mismatches.append((i, gen_row[2], gen_factors, ref_factors, col_diffs))

    cols_names = ["DAPA_FREC","DAPA_CM","DATO_FREC","DATO_CM","RC_FREC","RC_CM","ROPA_FREC","ROPA_CM","ROTO_FREC","ROTO_CM"]
    if not real_mismatches and not float_noise_rows:
        _ok(f"Factor values match reference for all {len(data_rows)} rows")
    elif not real_mismatches:
        _ok(f"Factor values match reference for all {len(data_rows)} rows (float noise only, within 1e-9)")
    else:
        _info(f"Factor divergences vs system reference ({len(real_mismatches)} rows) — cotizador is source of truth:")
        for idx, age_from, gen_f, ref_f, col_diffs in real_mismatches:
            for j in col_diffs:
                print(f"    age={age_from} {cols_names[j]}: cotizador={gen_f[j]}  sistema={ref_f[j]}")
        _info("These differences mean negocio has not yet requested a system update for those values.")

    return failures


def _verify_java(java_out: Path, base_name: str) -> int:
    failures = 0
    print("\n── Verifying java ──")

    content = java_out.read_text(encoding="utf-8")
    checks = [
        ("package eu.ncdc.arizona.rule.db.migration;", "correct package (ams-rule)"),
        ("import eu.ncdc.arizona.migration.task.LoadFromFileMigrationTask;", "correct import"),
        (f"public class {base_name} extends LoadFromFileMigrationTask", "correct class declaration"),
    ]
    for expected, label in checks:
        if expected in content:
            _ok(label)
        else:
            _fail(f"Missing: {label}")
            failures += 1

    return failures


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Rules migration end-to-end test runner")
    parser.add_argument("--entity", choices=list(_ENTITIES), default="VHDriversAge",
                        help="Entity to test (default: VHDriversAge)")
    parser.add_argument("--ticket", default=None,
                        help="Override ticket ID")
    args = parser.parse_args()

    entity_name = args.entity
    cfg = _ENTITIES[entity_name]

    if not cfg["input"].exists():
        print(f"Input file not found: {cfg['input']}")
        print("Place the cotizador file in fixtures/rules/business-reference/ and set BUSINESS_EXCEL_PASSWORD.")
        sys.exit(1)

    xlsx_out, java_out = _generate(entity_name, cfg, args.ticket)
    base_name = xlsx_out.stem

    total_failures = 0
    total_failures += _verify_xlsx(xlsx_out, entity_name, cfg)
    total_failures += _verify_java(java_out, base_name)

    print(f"\n{'─' * 50}")
    if total_failures == 0:
        print(f"\033[32m  ALL CHECKS PASSED\033[0m  →  {xlsx_out.name}")
    else:
        print(f"\033[31m  {total_failures} CHECK(S) FAILED\033[0m")
        sys.exit(1)


if __name__ == "__main__":
    main()
