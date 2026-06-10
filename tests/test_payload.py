"""Tests for run-payload mode (JSON payload → migration generation)."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

# project root on path
sys.path.insert(0, str(Path(__file__).parent.parent))

import src.config as _config_module
from src.description import build_description


# ── helpers ───────────────────────────────────────────────────────────────────

def _make_negocio_xlsx(path: Path, rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CHASIS", "TASA FINAL", "PLACAS"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def _write_config(cfg_path: Path, repo: str) -> None:
    cfg_path.write_text(json.dumps({"repo": repo}), encoding="utf-8")


# ── tests: description derivation (integration with payload fields) ───────────

class TestPayloadDescriptionDerivation:
    def test_ren_data_description_from_payload_fields(self) -> None:
        desc = build_description("ren-data", month=8, year=2026)
        assert desc == "VH_ren_data_ago_2026"

    def test_rules_description_from_entity(self) -> None:
        desc = build_description("rules", entity="VHPlanRules")
        assert desc == "VHPlanRules"


# ── tests: config loading ─────────────────────────────────────────────────────

class TestConfigLoading:
    def test_missing_config_raises_file_not_found(self, tmp_path: Path) -> None:
        missing = tmp_path / "nonexistent_config.json"
        with patch.object(_config_module, "_CONFIG_PATH", missing):
            with pytest.raises(FileNotFoundError, match="config.json"):
                _config_module.load_config()

    def test_valid_config_returns_dict(self, tmp_path: Path) -> None:
        cfg_file = tmp_path / "config.json"
        _write_config(cfg_file, "../some-repo")
        with patch.object(_config_module, "_CONFIG_PATH", cfg_file):
            cfg = _config_module.load_config()
        assert cfg["repo"] == "../some-repo"


# ── tests: run-payload end-to-end ─────────────────────────────────────────────

class TestRunPayload:
    def test_ren_data_payload_generates_xlsx(self, tmp_path: Path) -> None:
        """Valid ren-data payload → xlsx generated without error."""
        from src.generator_ren_data import generate as gen_xlsx

        input_file = tmp_path / "baseticket.xlsx"
        _make_negocio_xlsx(input_file, [
            ("C001", 0.02,  "P001"),
            ("C002", 0.015, "P002"),
        ])
        output_file = tmp_path / "out.xlsx"
        gen_xlsx(input_file, output_file, year=2026, month=8)

        wb = openpyxl.load_workbook(output_file, data_only=True)
        assert "FixedRenewalData" in wb.sheetnames
        assert wb["FixedRenewalData"].max_row == 3  # 1 header + 2 rows

    def test_payload_unknown_command_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="Unknown command"):
            build_description("infra-change", month=8, year=2026)

    def test_ren_data_payload_description_in_filename(self, tmp_path: Path) -> None:
        """Description derived from month=8, year=2026 → 'VH_ren_data_ago_2026'."""
        desc = build_description("ren-data", month=8, year=2026)
        assert "ago" in desc
        assert "2026" in desc

    def test_rules_payload_description_matches_entity(self, tmp_path: Path) -> None:
        desc = build_description("rules", entity="VHPlanRules")
        assert desc == "VHPlanRules"
