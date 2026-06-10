"""
Migration runner for manual end-to-end testing.

Generates real migration files into tests/migrations/ and verifies them
against the reference Excel from the backend repo.

Usage:
    python tests/run_migration_test.py
    python tests/run_migration_test.py --month julio
    python tests/run_migration_test.py --month julio --ticket INC23703493
"""

from __future__ import annotations

import argparse
import sys
import tempfile
from datetime import datetime
from pathlib import Path

# ensure project root is on the path regardless of cwd
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

_ROOT = Path(__file__).parent.parent
_OUTPUT_DIR = Path(__file__).parent / "migrations"
_REQUIREMENTS = _ROOT / "requirements" / "renovaciones" / "2026"
_BACKEND = _ROOT.parent / "ov-arizona-backend-ecuador"

# Known month configs
_MONTHS = {
    "abril": {
        "file":   _REQUIREMENTS / "abril" / "baseticketAbril2026.xlsx",
        "year":   2026,
        "month":  4,
        "ticket": "INC23540797",
        "desc":   "VH_ren_data_abr",
    },
    "mayo": {
        "file":   _REQUIREMENTS / "mayo" / "baseticketMayo2026.xlsx",
        "year":   2026,
        "month":  5,
        "ticket": "INC23649033",
        "desc":   "VH_ren_data_may",
    },
    "junio": {
        "file":   _REQUIREMENTS / "junio" / "baseticketJunio2026.xlsx",
        "year":   2026,
        "month":  6,
        "ticket": "INC23703493",
        "desc":   "VH_ren_data_jun",
    },
    "julio": {
        "file":   _REQUIREMENTS / "julio" / "baseticketJulio2026.xlsx",
        "year":   2026,
        "month":  7,
        "ticket": "INC23703493",
        "desc":   "VH_ren_data_jul",
        "reference_xlsx": _BACKEND / "ams-policy/flyway/src/main/resources/db/migration"
                          / "V2026_05_15_12_00_00__INC23703493_VH_ren_data_jul.xlsx",
    },
}

# ── helpers ───────────────────────────────────────────────────────────────────

def _ok(msg: str) -> None:
    print(f"  \033[32m✓\033[0m {msg}")

def _fail(msg: str) -> None:
    print(f"  \033[31m✗\033[0m {msg}")

def _info(msg: str) -> None:
    print(f"  · {msg}")


# ── generation ────────────────────────────────────────────────────────────────

def _generate(cfg: dict, ticket_override: str | None) -> tuple[Path, Path]:
    """Generate xlsx + java into tests/migrations/ and return their paths."""
    from src.generator_ren_data import generate as gen_xlsx
    from src.java_template import generate as gen_java

    ticket = ticket_override or cfg["ticket"]
    ts = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    base_name = f"V{ts}__{ticket}_{cfg['desc']}"

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_out = _OUTPUT_DIR / f"{base_name}.xlsx"
    java_out = _OUTPUT_DIR / f"{base_name}.java"

    print(f"\n── Generating {base_name} ──")
    gen_xlsx(cfg["file"], xlsx_out, year=cfg["year"], month=cfg["month"])

    java_src = gen_java(base_name, "ams-policy")
    java_out.write_text(java_src, encoding="utf-8")
    _info(f"Java: {java_out.name}")

    return xlsx_out, java_out


# ── verification ──────────────────────────────────────────────────────────────

def _verify_xlsx(xlsx_out: Path, cfg: dict) -> int:
    """Run assertions on the generated xlsx. Returns number of failures."""
    failures = 0

    print("\n── Verifying xlsx ──")

    wb = openpyxl.load_workbook(xlsx_out, data_only=True)

    # 1. Sheet names and order
    if wb.sheetnames == ["LOV", "FixedRenewalData"]:
        _ok(f"Sheets: {wb.sheetnames}")
    else:
        _fail(f"Expected ['LOV', 'FixedRenewalData'], got {wb.sheetnames}")
        failures += 1

    ws = wb["FixedRenewalData"]

    # 2. Headers
    expected_headers = [
        "FixedRenewalData", "Year", "Month", "Chassis number",
        "Sum insured", "Factor", "Renewal blocked",
    ]
    actual_headers = [ws.cell(1, c).value for c in range(1, 8)]
    if actual_headers == expected_headers:
        _ok(f"Headers: {actual_headers}")
    else:
        _fail(f"Headers mismatch.\n    Expected: {expected_headers}\n    Got:      {actual_headers}")
        failures += 1

    # 3. Row count
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    _info(f"Data rows: {len(data_rows)}")

    # 4. Year / Month injected
    years  = {r[1] for r in data_rows}
    months = {r[2] for r in data_rows}
    if years == {cfg["year"]}:
        _ok(f"Year = {cfg['year']} in all rows")
    else:
        _fail(f"Unexpected year values: {years}")
        failures += 1
    if months == {cfg["month"]}:
        _ok(f"Month = {cfg['month']} in all rows")
    else:
        _fail(f"Unexpected month values: {months}")
        failures += 1

    # 5. ID column always None
    id_nulls = all(r[0] is None for r in data_rows)
    if id_nulls:
        _ok("ID column (col A) is None in all rows")
    else:
        _fail("ID column has non-None values")
        failures += 1

    # 6. No Renovar → Renewal blocked = Yes
    no_renovar = [r for r in data_rows if isinstance(r[5], str) and r[5].strip().lower() == "no renovar"]
    blocked_yes = [r for r in no_renovar if r[6] == "Yes"]
    if no_renovar:
        if len(no_renovar) == len(blocked_yes):
            _ok(f"No Renovar rows: {len(no_renovar)} — all have Renewal blocked=Yes")
        else:
            _fail(f"{len(no_renovar) - len(blocked_yes)} No Renovar rows missing Renewal blocked=Yes")
            failures += 1
    else:
        _info("No 'No Renovar' rows in this file")

    # 7. Normal rows → Renewal blocked = No
    normal = [r for r in data_rows if isinstance(r[5], (int, float))]
    blocked_no = [r for r in normal if r[6] == "No"]
    if len(normal) == len(blocked_no):
        _ok(f"Normal rows: {len(normal)} — all have Renewal blocked=No")
    else:
        _fail(f"{len(normal) - len(blocked_no)} normal rows with unexpected Renewal blocked value")
        failures += 1

    # 8. LOV row count
    ws_lov = wb["LOV"]
    lov_rows = [r for r in ws_lov.iter_rows(values_only=True) if any(c is not None for c in r)]
    if len(lov_rows) == 289:
        _ok(f"LOV rows: {len(lov_rows)}")
    else:
        _fail(f"LOV expected 289 rows, got {len(lov_rows)}")
        failures += 1

    # 9. Compare against reference if available
    ref_path: Path | None = cfg.get("reference_xlsx")
    if ref_path and ref_path.exists():
        print("\n── Comparing against reference ──")
        failures += _compare_with_reference(ws, data_rows, ref_path)
    elif ref_path:
        _info(f"Reference not found, skipping comparison: {ref_path.name}")

    return failures


def _compare_with_reference(ws_out, data_rows: list, ref_path: Path) -> int:
    failures = 0
    wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
    ws_ref = wb_ref["FixedRenewalData"]
    ref_rows = [r for r in ws_ref.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

    # Row count
    if len(data_rows) == len(ref_rows):
        _ok(f"Row count matches reference: {len(data_rows)}")
    else:
        _fail(f"Row count: generated={len(data_rows)}, reference={len(ref_rows)}")
        failures += 1

    # Chassis sets
    gen_chassis  = {r[3] for r in data_rows}
    ref_chassis  = {r[3] for r in ref_rows}
    only_in_gen  = gen_chassis - ref_chassis
    only_in_ref  = ref_chassis - gen_chassis

    if not only_in_gen and not only_in_ref:
        _ok(f"Chassis set matches reference ({len(gen_chassis)} unique)")
    else:
        if only_in_gen:
            _fail(f"Chassis in generated but NOT in reference ({len(only_in_gen)}): {list(only_in_gen)[:5]}")
            failures += 1
        if only_in_ref:
            _fail(f"Chassis in reference but NOT in generated ({len(only_in_ref)}): {list(only_in_ref)[:5]}")
            failures += 1

    # Factor values per chassis
    gen_factor = {r[3]: r[5] for r in data_rows}
    ref_factor = {r[3]: r[5] for r in ref_rows}
    mismatches = [
        (c, gen_factor[c], ref_factor[c])
        for c in gen_chassis & ref_chassis
        if gen_factor[c] != ref_factor[c]
    ]
    if not mismatches:
        _ok("Factor values match reference for all shared chassis")
    else:
        _fail(f"Factor mismatches ({len(mismatches)}):")
        for chassis, gen_v, ref_v in mismatches[:5]:
            print(f"    {chassis}: generated={gen_v}, reference={ref_v}")
        failures += 1

    # No Renovar count
    gen_nr = sum(1 for r in data_rows if isinstance(r[5], str) and r[5].lower() == "no renovar")
    ref_nr = sum(1 for r in ref_rows  if isinstance(r[5], str) and r[5].lower() == "no renovar")
    if gen_nr == ref_nr:
        _ok(f"No Renovar count matches reference: {gen_nr}")
    else:
        _fail(f"No Renovar: generated={gen_nr}, reference={ref_nr}")
        failures += 1

    return failures


def _verify_java(java_out: Path, base_name: str) -> int:
    failures = 0
    print("\n── Verifying java ──")

    content = java_out.read_text(encoding="utf-8")
    checks = [
        ("package eu.ncdc.arizona.policy.db.migration;", "correct package"),
        ("import eu.ncdc.arizona.migration.task.LoadFromFileMigrationTask;", "correct import"),
        (f"public class {base_name} extends LoadFromFileMigrationTask", "correct class declaration"),
    ]
    for expected, label in checks:
        if expected in content:
            _ok(label)
        else:
            _fail(f"Missing: {label}")
            failures += 1

    return failures


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Migration end-to-end test runner")
    parser.add_argument("--month", choices=list(_MONTHS), default="julio",
                        help="Month to test (default: julio)")
    parser.add_argument("--ticket", default=None,
                        help="Override ticket ID")
    args = parser.parse_args()

    cfg = _MONTHS[args.month]

    if not cfg["file"].exists():
        print(f"Input file not found: {cfg['file']}")
        sys.exit(1)

    xlsx_out, java_out = _generate(cfg, args.ticket)
    base_name = xlsx_out.stem

    total_failures = 0
    total_failures += _verify_xlsx(xlsx_out, cfg)
    total_failures += _verify_java(java_out, base_name)

    print(f"\n{'─' * 50}")
    if total_failures == 0:
        print(f"\033[32m  ALL CHECKS PASSED\033[0m  →  {xlsx_out.name}")
    else:
        print(f"\033[31m  {total_failures} CHECK(S) FAILED\033[0m")
        sys.exit(1)


if __name__ == "__main__":
    main()
