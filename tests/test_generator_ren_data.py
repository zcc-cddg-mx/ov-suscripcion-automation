"""Tests for generator_ren_data."""

from pathlib import Path

import openpyxl
import pytest

from src.generator_ren_data import generate, _validate_and_normalize_factor, _decimal_places, _MAX_FACTOR_DECIMALS

_YEAR = 2026
_MONTH = 8


def _make_raw_input_full(path: Path, rows: list[dict]) -> None:
    """Input with year/month columns (legacy / test-only format)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["chassis", "year", "month", "factor", "sum insured", "renewal blocked"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(path)


def _make_raw_input_negocio(path: Path, rows: list[tuple]) -> None:
    """Simulates the real format from negocio: CHASIS, TASA FINAL, PLACAS + empty trailing rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CHASIS", "TASA FINAL", "PLACAS", None])
    for r in rows:
        ws.append(list(r) + [None])
    # add empty trailing rows (as in real files)
    for _ in range(3):
        ws.append([None, None, None, None])
    wb.save(path)


class TestGeneratorRenData:
    def test_basic_output_structure(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_full(raw, [
            {"chassis": "ABC123", "year": 2026, "month": 8, "factor": 0.02, "sum insured": None, "renewal blocked": "No"},
            {"chassis": "XYZ999", "year": 2026, "month": 8, "factor": 0.015, "sum insured": 10000, "renewal blocked": "No"},
        ])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        assert out.exists()
        wb = openpyxl.load_workbook(out, data_only=True)
        assert wb.sheetnames == ["LOV", "FixedRenewalData"]

    def test_lov_sheet_has_289_rows(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [("TEST001", 0.019, "AAA111")])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws_lov = wb["LOV"]
        rows = [r for r in ws_lov.iter_rows(values_only=True) if any(c is not None for c in r)]
        assert len(rows) == 289

    def test_fixed_renewal_data_headers(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [("TEST001", 0.019, "AAA111")])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        headers = [ws.cell(1, col).value for col in range(1, 8)]
        assert headers == [
            "FixedRenewalData", "Year", "Month",
            "Chassis number", "Sum insured", "Factor", "Renewal blocked",
        ]

    def test_negocio_format_injects_year_month(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [("CHASSIS001", 0.025, "BBB222")])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=2026, month=9)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        assert ws.cell(2, 2).value == 2026  # Year
        assert ws.cell(2, 3).value == 9     # Month

    def test_negocio_format_filters_empty_rows(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [
            ("C001", 0.02, "P001"),
            ("C002", 0.03, "P002"),
        ])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        assert ws.max_row == 3  # 1 header + 2 data rows (no empty rows)

    def test_no_renovar_rows_are_included_with_renewal_blocked(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [
            ("C001", 0.02,          "P001"),
            ("C002", "No Renovar",  "P002"),
            ("C003", "NO RENOVAR",  "P003"),
            ("C004", 0.019,         "P004"),
        ])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        # All 4 rows must be present
        assert ws.max_row == 5  # 1 header + 4 data rows

        # No Renovar rows: Factor preserved as string, Renewal blocked = Yes
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        no_renovar_rows = [r for r in rows if isinstance(r[5], str) and r[5].lower() == "no renovar"]
        assert len(no_renovar_rows) == 2
        for r in no_renovar_rows:
            assert r[6] == "Yes"  # Renewal blocked

        # Normal rows: Renewal blocked = No
        normal_rows = [r for r in rows if isinstance(r[5], float)]
        assert len(normal_rows) == 2
        for r in normal_rows:
            assert r[6] == "No"

    def test_id_column_is_none(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [("CHASSIS001", 0.02, "P001")])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        assert ws.cell(2, 1).value is None

    def test_missing_required_column_raises(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["year", "month"])  # missing chassis and factor
        ws.append([2026, 8])
        raw = tmp_path / "bad.xlsx"
        wb.save(raw)
        out = tmp_path / "output.xlsx"
        with pytest.raises(ValueError, match="Missing required columns"):
            generate(raw, out, year=_YEAR, month=_MONTH)

    def test_all_no_renovar_still_generates(self, tmp_path: Path) -> None:
        """A file with only No Renovar rows is valid — all are included with Renewal blocked=Yes."""
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [
            ("C001", "No Renovar", "P001"),
            ("C002", "No Renovar", "P002"),
        ])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        assert ws.max_row == 3  # 1 header + 2 rows
        assert ws.cell(2, 7).value == "Yes"
        assert ws.cell(3, 7).value == "Yes"

    def test_accepts_tasa_final_column_name(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw.xlsx"
        _make_raw_input_negocio(raw, [("ABC123", 0.02, "PLACA1")])
        out = tmp_path / "output.xlsx"
        generate(raw, out, year=_YEAR, month=_MONTH)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        assert ws.cell(2, 6).value == 0.02  # Factor col

    def test_invalid_factor_type_raises(self, tmp_path: Path) -> None:
        """A non-numeric, non-'No Renovar' Factor value must raise ValueError."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["CHASIS", "TASA FINAL", "PLACAS"])
        ws.append(["C001", "INVALIDO", "P001"])
        raw = tmp_path / "bad_factor.xlsx"
        wb.save(raw)
        out = tmp_path / "output.xlsx"
        with pytest.raises(ValueError, match="Factor must be numeric or 'No Renovar'"):
            generate(raw, out, year=_YEAR, month=_MONTH)

    def test_real_requirements_file_julio(self, tmp_path: Path) -> None:
        """Smoke test against real julio file: 1589 rows total, 14 No Renovar included."""
        real_file = Path(__file__).parent.parent / "requirements/renovaciones/2026/julio/baseticketJulio2026.xlsx"
        if not real_file.exists():
            pytest.skip("Real requirements file not available")

        out = tmp_path / "output.xlsx"
        generate(real_file, out, year=2026, month=7)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["FixedRenewalData"]
        assert ws.max_row == 1590  # 1 header + 1589 data rows (no row excluded)
        assert ws.cell(2, 2).value == 2026
        assert ws.cell(2, 3).value == 7

        # Verify 14 No Renovar rows are included with Renewal blocked=Yes
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        blocked = [r for r in rows if r[6] == "Yes"]
        assert len(blocked) == 14
        for r in blocked:
            assert isinstance(r[5], str) and r[5].lower() == "no renovar"

        # All numeric factors must have at most 8 decimal places
        numeric_rows = [r for r in rows if isinstance(r[5], float)]
        for r in numeric_rows:
            assert _decimal_places(r[5]) <= _MAX_FACTOR_DECIMALS, \
                f"Factor {r[5]} has more than {_MAX_FACTOR_DECIMALS} decimals"


class TestFactorValidation:
    def test_valid_float_under_8_decimals_unchanged(self) -> None:
        assert _validate_and_normalize_factor(0.019, "C001", 2) == 0.019

    def test_valid_float_exactly_8_decimals_unchanged(self) -> None:
        assert _validate_and_normalize_factor(0.01963615, "C001", 2) == 0.01963615

    def test_float_over_8_decimals_is_rounded(self) -> None:
        result = _validate_and_normalize_factor(0.0396510705789056, "C001", 2)
        assert result == round(0.0396510705789056, 8)
        assert _decimal_places(result) <= _MAX_FACTOR_DECIMALS

    def test_int_factor_converted_to_float(self) -> None:
        result = _validate_and_normalize_factor(1, "C001", 2)
        assert result == 1.0
        assert isinstance(result, float)

    def test_no_renovar_passthrough(self) -> None:
        assert _validate_and_normalize_factor("No Renovar", "C001", 2) == "No Renovar"

    def test_no_renovar_case_insensitive(self) -> None:
        assert _validate_and_normalize_factor("NO RENOVAR", "C001", 2) == "NO RENOVAR"
        assert _validate_and_normalize_factor("no renovar", "C001", 2) == "no renovar"

    def test_arbitrary_string_raises(self) -> None:
        with pytest.raises(ValueError, match="Factor must be numeric or 'No Renovar'"):
            _validate_and_normalize_factor("INVALIDO", "C001", 2)

    def test_none_raises(self) -> None:
        with pytest.raises(ValueError, match="Factor must be numeric or 'No Renovar'"):
            _validate_and_normalize_factor(None, "C001", 2)

    def test_error_message_includes_chassis_and_row(self) -> None:
        with pytest.raises(ValueError, match="chassis 'CHASSIS_X'") as exc:
            _validate_and_normalize_factor("MAL", "CHASSIS_X", 42)
        assert "Row 42" in str(exc.value)

    def test_decimal_places_helper(self) -> None:
        assert _decimal_places(0.02) == 2
        assert _decimal_places(0.01963615) == 8
        assert _decimal_places(0.0396510705789056) > 8
